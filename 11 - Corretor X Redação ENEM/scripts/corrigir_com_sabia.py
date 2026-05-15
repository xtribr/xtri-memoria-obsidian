#!/usr/bin/env python3
"""Corrige uma redação ENEM com Sabiá e preenche o template Excel do Corretor X.

Uso mínimo:
  export SABIA_API_KEY="..."
  python3 scripts/corrigir_com_sabia.py \
    --tema-file entradas/caso-001/tema.txt \
    --redacao-file entradas/caso-001/redacao.txt \
    --case-id CASO-001 \
    --aluno-id "Aluno 001" \
    --out-xlsx "Cérebro do 1000/casos/exports/CASO-001.xlsx"
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import time
import unicodedata
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from schemas.correcao import (
    GateAnulacao as GateAnulacaoSchema,
    MetadadosRedacao,
    RespostaC1,
    RespostaC2,
    RespostaC3,
    RespostaC4,
    RespostaC5,
    ResultadoCorrecao as ResultadoCorrecaoSchema,
    parse_resposta_sabia,
    schema_for_etapa,
)


API_URL = "https://chat.maritaca.ai/api/chat/completions"
DEFAULT_MODEL = "sabia-4"
DEFAULT_PROMPT_VERSION = "v1.0"
DEFAULT_FALLBACK_MODEL = os.environ.get("SABIA_FALLBACK_MODEL", DEFAULT_MODEL)
REVISAO_HUMANA_EXIT_CODE = 3
RETRY_POLICY = {
    "tentativas_max": 2,
    "estrategias": [
        "re_prompt_com_erro",
        "fallback_modelo_maior",
        "marcar_revisao_humana",
    ],
}
VALID_SCORES = {0, 40, 80, 120, 160, 200}
COMPETENCIAS = ["C1", "C2", "C3", "C4", "C5"]
ORDEM_CORRECAO = ["C2", "C1", "C3", "C4", "C5"]
COMPETENCIAS_COM_TETO_TANGENCIAMENTO = {"C3", "C5"}
DEVOLUTIVA_HEADERS = [
    "id_redacao",
    "data_correcao",
    "aluno_nome",
    "aluno_escola",
    "tema",
    "status_tema",
    "anulada",
    "motivos_anulacao",
    "tangenciamento",
    "nota_c1",
    "nota_c2",
    "nota_c3",
    "nota_c4",
    "nota_c5",
    "nota_final",
    "teto_aplicado_c3",
    "teto_aplicado_c5",
    "comentario_c1",
    "comentario_c2",
    "comentario_c3",
    "comentario_c4",
    "comentario_c5",
    "sugestao_c1",
    "sugestao_c2",
    "sugestao_c3",
    "sugestao_c4",
    "sugestao_c5",
    "confianca_geral",
    "alertas",
]
RESUMO_HEADERS = [
    "id_redacao",
    "data_correcao",
    "aluno_escola",
    "tema",
    "nota_final",
    "nivel_proficiencia",
    "nota_c1",
    "nota_c2",
    "nota_c3",
    "nota_c4",
    "nota_c5",
    "total_desvios_c1",
    "reincidencia_c1",
    "abordagem_tema_c2",
    "qtd_repertorios",
    "qtd_repertorios_bolso",
    "projeto_texto_c3",
    "autoria_c3",
    "diversidade_coesiva_c4",
    "total_elementos_c5",
    "articulacao_proposta_c5",
    "respeita_dh_c5",
    "confianca_geral",
    "custo_tokens_total",
    "tempo_processamento_s",
]
CALIBRACAO_HEADERS = [
    "id_redacao",
    "competencia",
    "tipo_apontamento",
    "subtipo",
    "trecho_evidencia",
    "gravidade",
    "classificacao",
    "justificativa",
]
AUDITORIA_HEADERS = [
    "id_chamada",
    "id_redacao",
    "timestamp",
    "etapa",
    "modelo",
    "prompt_versao",
    "prompt_hash",
    "json_resposta_bruto",
    "validacao_pydantic_ok",
    "erros_pydantic",
    "tempo_resposta_s",
    "tokens_input",
    "tokens_output",
    "custo_estimado_usd",
]
REPERTORIO_SUBTIPO_ALIASES = {
    "pequeno manual antirracista": "djamila_ribeiro",
}


RUBRICAS = {
    "C1": (
        "Competência I: demonstrar domínio da modalidade escrita formal da língua portuguesa. "
        "Avalie desvios gramaticais, convenções da escrita, registro, vocabulário e estrutura sintática."
    ),
    "C2": (
        "Competência II: compreender a proposta de redação e desenvolver o tema nos limites "
        "do texto dissertativo-argumentativo. Avalie aderência ao tema, tipo textual e repertório "
        "sociocultural legitimado, pertinente e produtivo."
    ),
    "C3": (
        "Competência III: selecionar, relacionar, organizar e interpretar informações, fatos, "
        "opiniões e argumentos em defesa de um ponto de vista. Avalie projeto de texto, autoria, "
        "progressão e consistência argumentativa."
    ),
    "C4": (
        "Competência IV: demonstrar conhecimento dos mecanismos linguísticos necessários para "
        "a construção da argumentação. Avalie coesão, operadores argumentativos, referenciação "
        "e articulação entre períodos e parágrafos."
    ),
    "C5": (
        "Competência V: elaborar proposta de intervenção para o problema abordado, respeitando "
        "os direitos humanos. Avalie agente, ação, meio, efeito/finalidade, detalhamento e "
        "articulação com a discussão."
    ),
}

PROMPT_BASE = """Você é o Corretor X da XTRI EdTECH, especialista em redação ENEM e Teoria de Resposta ao Item.

TAREFA
Avalie EXCLUSIVAMENTE a {competencia} da redação abaixo, atribuindo uma das seis notas oficiais: 0, 40, 80, 120, 160 ou 200.

AUTORIDADE METODOLÓGICA (ordem de precedência)
1. Cartilha do Participante ENEM 2025 (INEP/MEC) — referência principal e final
2. Matriz de Referência para a Redação do ENEM
3. Redações nota 1000 de edições anteriores — referência de padrão, NUNCA regra absoluta

GATES DE ANULAÇÃO (verificar ANTES de avaliar)
Se o campo "status_anulacao" abaixo indicar qualquer condição abaixo, atribua nota 0 a esta competência e explique no campo "comentario":
- fuga_total_tema
- nao_atendimento_tipo_textual
- texto_insuficiente (≤7 linhas manuscritas)
- parte_deliberadamente_desconectada
- texto_em_lingua_estrangeira
- texto_ilegivel
- improperios_ou_anulacao_proposital

TANGENCIAMENTO (aplicável a C3 e C5)
Se "tangenciamento_c2" = true, a nota desta competência não pode exceder 40 pontos, conforme Matriz de Referência ENEM.

REGRAS OBRIGATÓRIAS
- Avalie SOMENTE a competência indicada. Ignore problemas que pertençam a outras competências.
- Cite evidências REAIS do texto do aluno (trechos curtos, entre aspas).
- NUNCA invente tema, erro, repertório, autor, dado estatístico ou citação.
- Se faltar informação para avaliação segura, reduza "nivel_confianca" para "media" ou "baixa".
- Não reescreva o texto do aluno nem produza redação modelo.
- Rigor pedagógico, sem condescendência e sem agressividade.
- Para 200 pontos, o texto precisa atender a TODOS os critérios da faixa, não apenas alguns.
- Em caso de dúvida entre duas faixas, escolha a INFERIOR e justifique.

REGRA DE TRANSCRIÇÃO FORENSE
- A redação abaixo deve ser tratada como transcrição literal do manuscrito.
- Não normalize, corrija ou "melhore" grafia, acentos, pontuação, concordância, paragrafação ou escolhas vocabulares do aluno.
- Preserve a interpretação de erros aparentes como possíveis erros reais do aluno, salvo quando o status OCR indicar baixa confiança ou trecho incerto.
- Se o status OCR/transcrição não indicar validação segura, reduza "nivel_confianca" e explicite a limitação no comentário.
- Para C1, use desvios presentes na transcrição literal como evidências, sem supor correção automática.

CONTEXTO DA AVALIAÇÃO
- Tema oficial: {tema}
- Status do tema: {status_tema}  # verificado | inferido | ausente
- Status OCR/transcrição: {status_ocr}
- Status de anulação: {status_anulacao}
- Tangenciamento detectado em C2: {tangenciamento_c2}

RUBRICA DESTA COMPETÊNCIA
{rubrica}

REDAÇÃO DO ALUNO
{redacao}

FORMATO DE SAÍDA (JSON válido, sem markdown, sem texto fora do JSON)
{formato_json}
"""

PROMPT_GATE_ANULACAO = """Você é o Corretor X da XTRI EdTECH. Sua tarefa é verificar se a redação abaixo apresenta alguma condição de anulação total, ANTES da avaliação por competência.

CONDIÇÕES DE ANULAÇÃO (Cartilha ENEM 2025, p. 7-8)
1. fuga_total_tema: nem o assunto amplo é abordado
2. nao_atendimento_tipo_textual: texto não é dissertativo-argumentativo (é poema, narração pura, carta, lista, etc.)
3. texto_insuficiente: até 7 linhas manuscritas (ou 10 em Braille)
4. parte_deliberadamente_desconectada: bilhetes para a banca, orações, mensagens políticas/religiosas desarticuladas, trechos de música/poema sem função argumentativa
5. texto_em_lingua_estrangeira: predominantemente ou integralmente em outra língua
6. texto_ilegivel: impossível de ser lido
7. improperios_ou_anulacao_proposital: ofensas, desenhos, formas propositais de anulação

REGRAS
- Avalie cada condição independentemente
- Se NENHUMA for true, retorne anulado = false e prossiga normalmente
- Se ALGUMA for true, retorne anulado = true e indique qual
- Para tangenciamento (não é anulação total), avalie em C2

CONTEXTO
- Tema oficial: {tema}
- Status do tema: {status_tema}
- Status OCR: {status_ocr}
- Número de linhas estimadas: {num_linhas}

REDAÇÃO DO ALUNO
{redacao}

FORMATO DE SAÍDA (JSON válido, sem markdown)
{{
  "anulado": false,
  "motivos": [],
  "detalhes": {{
    "fuga_total_tema": false,
    "nao_atendimento_tipo_textual": false,
    "texto_insuficiente": false,
    "parte_deliberadamente_desconectada": false,
    "texto_em_lingua_estrangeira": false,
    "texto_ilegivel": false,
    "improperios_ou_anulacao_proposital": false
  }},
  "evidencia": "trecho ou descrição que justifica a anulação, se houver",
  "nivel_confianca": "alta|media|baixa"
}}
"""

FORMATO_JSON = """{{
  "competencia": "{competencia}",
  "nota_corretor_x": 0,
  "comentario_do_erro": "comentário pedagógico objetivo",
  "evidencia_no_texto": "trecho curto do aluno entre aspas, ou descrição fiel se o trecho estiver ilegível",
  "sugestao_de_melhoria": "orientação prática sem substituir a autoria do aluno",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_C1 = """{{
  "competencia": "C1",
  "nota": 0,
  "comentario": "análise pedagógica objetiva (3-5 frases)",
  "desvios_encontrados": [
    {{"tipo": "ortografia|acentuacao|pontuacao|concordancia|regencia|crase|paralelismo|registro|vocabulario|sintaxe", "trecho": "citação exata", "gravidade": "leve|media|grave"}}
  ],
  "total_desvios": 0,
  "reincidencia": false,
  "tipo_reincidente": null,
  "sugestao": "orientação prática focada no desvio mais impactante",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_C2 = """{{
  "competencia": "C2",
  "nota": 0,
  "comentario": "análise pedagógica objetiva (3-5 frases)",
  "abordagem_tema": "completa|tangenciamento|fuga_total",
  "tangenciamento_c2": false,
  "alerta_tema": null,
  "tipo_textual_adequado": true,
  "repertorios_identificados": [
    {{"referencia": "ex: Djamila Ribeiro - Pequeno Manual Antirracista", "classificacao": "produtivo|bolso|nao_legitimado", "justificativa": "breve"}}
  ],
  "sugestao": "orientação prática",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_C3 = """{{
  "competencia": "C3",
  "nota": 0,
  "comentario": "análise pedagógica objetiva (3-5 frases)",
  "projeto_de_texto": "consistente|com_indicios|previsivel|ausente",
  "tese_identificada": "citação curta da tese, se houver",
  "autoria": "configura|indicios|ausente",
  "teto_por_tangenciamento_aplicado": false,
  "sugestao": "orientação prática",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_C4 = """{{
  "competencia": "C4",
  "nota": 0,
  "comentario": "análise pedagógica objetiva (3-5 frases)",
  "articulacao_interparagrafo": "plena|consistente|mediana|insuficiente|precaria|ausente",
  "diversidade_coesiva": "alta|media|baixa",
  "inadequacoes_identificadas": [
    {{"tipo": "conector_sem_relacao|empilhamento|repeticao_inadequada|paragrafo_isolado", "trecho": "citação", "justificativa": "breve"}}
  ],
  "sugestao": "orientação prática",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_C5 = """{{
  "competencia": "C5",
  "nota": 0,
  "comentario": "análise pedagógica objetiva (3-5 frases)",
  "elementos_identificados": {{
    "agente": "citação ou null",
    "acao": "citação ou null",
    "meio": "citação ou null",
    "efeito": "citação ou null",
    "detalhamento": "citação ou null"
  }},
  "total_elementos": 0,
  "articulacao_com_texto": true,
  "respeita_direitos_humanos": true,
  "teto_por_tangenciamento_aplicado": false,
  "sugestao": "orientação prática",
  "nivel_confianca": "alta|media|baixa"
}}"""

FORMATO_JSON_POR_COMPETENCIA = {
    "C1": FORMATO_JSON_C1,
    "C2": FORMATO_JSON_C2,
    "C3": FORMATO_JSON_C3,
    "C4": FORMATO_JSON_C4,
    "C5": FORMATO_JSON_C5,
}


@dataclass
class AvaliacaoCompetencia:
    competencia: str
    nota_corretor_x: int
    comentario_do_erro: str
    evidencia_no_texto: str
    sugestao_de_melhoria: str
    nivel_confianca: str
    teto_tangenciamento_aplicado: bool = False


@dataclass
class GateAnulacao:
    anulado: bool
    motivos: list[str]
    evidencia: str
    nivel_confianca: str


@dataclass
class SabiaResponse:
    content: str
    response_body: dict[str, Any]
    tempo_resposta_s: float
    tokens_input: int | None = None
    tokens_output: int | None = None
    custo_estimado_usd: float | None = None


@dataclass
class ChamadaAuditoria:
    id_chamada: str
    timestamp: datetime
    etapa: str
    modelo: str
    prompt_versao: str
    prompt_hash: str
    json_resposta_bruto: str
    validacao_pydantic_ok: bool
    erros_pydantic: str
    tempo_resposta_s: float
    tokens_input: int | None = None
    tokens_output: int | None = None
    custo_estimado_usd: float | None = None


class SabiaValidationError(ValueError):
    """Erro lançado quando a resposta do Sabiá falha nos schemas Pydantic."""

    def __init__(
        self,
        etapa: str,
        erro: str,
        auditoria: ChamadaAuditoria | None = None,
        raw: dict[str, Any] | None = None,
    ) -> None:
        self.etapa = etapa
        self.erro = erro
        self.auditoria = auditoria
        self.raw = raw or {}
        super().__init__(f"Resposta inválida do Sabiá na etapa {etapa}: {erro}")


class SabiaHumanReviewRequired(RuntimeError):
    """Erro final quando a política de recuperação esgota as tentativas."""

    def __init__(self, etapa: str, erro: str, chamadas_auditoria: list[ChamadaAuditoria]) -> None:
        self.etapa = etapa
        self.erro = erro
        self.chamadas_auditoria = chamadas_auditoria
        super().__init__(f"Revisão humana recomendada na etapa {etapa}: {erro}")


@dataclass
class ResultadoCorrecao:
    anulado: bool
    motivos_anulacao: list[str]
    tangenciamento: bool
    nota_final: int
    avaliacoes: list[AvaliacaoCompetencia]
    raw_respostas: dict[str, dict[str, Any]]
    chamadas_auditoria: list[ChamadaAuditoria]
    resultado_pydantic: ResultadoCorrecaoSchema | None = None
    custo_tokens_total: int | None = None
    tempo_processamento_s: float | None = None


def read_text(path: Path) -> str:
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        raise ValueError(f"Arquivo vazio: {path}")
    return text


def read_text_or_default(path: Path, default: str) -> str:
    text = path.read_text(encoding="utf-8").strip()
    return text if text else default


def load_competencia_prompt(competencia: str, prompts_dir: Path) -> str:
    prompt_path = prompts_dir / f"{competencia.lower()}.md"
    if prompt_path.exists():
        return read_text(prompt_path)
    return RUBRICAS[competencia]


def extract_json(content: str) -> dict[str, Any]:
    content = content.strip()
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", content, re.DOTALL)
    if fenced:
        content = fenced.group(1)
    else:
        start = content.find("{")
        end = content.rfind("}")
        if start >= 0 and end > start:
            content = content[start : end + 1]
    return json.loads(content)


def prompt_hash(prompt: str) -> str:
    return hashlib.sha256(prompt[:500].encode("utf-8")).hexdigest()


def int_or_none(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def float_or_none(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def usage_value(usage: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        if key in usage:
            return usage[key]
    return None


def extract_sabia_usage(body: dict[str, Any]) -> tuple[int | None, int | None, float | None]:
    usage = body.get("usage")
    if not isinstance(usage, dict):
        return None, None, None

    tokens_input = int_or_none(
        usage_value(usage, ("prompt_tokens", "input_tokens", "tokens_input", "input"))
    )
    tokens_output = int_or_none(
        usage_value(usage, ("completion_tokens", "output_tokens", "tokens_output", "output"))
    )
    custo_estimado_usd = float_or_none(
        usage_value(usage, ("cost_usd", "estimated_cost_usd", "custo_estimado_usd", "cost"))
    )
    return tokens_input, tokens_output, custo_estimado_usd


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def infer_status_tema(tema: str) -> str:
    normalized = strip_accents(tema).lower()
    if any(marker in normalized for marker in ("nao informado", "ausente", "sem tema oficial")):
        return "ausente"
    if any(marker in normalized for marker in ("inferido", "inferencia", "aparenta tratar")):
        return "inferido"
    return "verificado"


def estimate_num_linhas(redacao: str) -> int:
    linhas = [linha for linha in redacao.splitlines() if linha.strip()]
    palavras = redacao.split()
    if not palavras:
        return 0
    estimativa_por_palavras = max(1, round(len(palavras) / 12))
    return max(len(linhas), estimativa_por_palavras)


def normalize_confidence(value: str) -> str:
    normalized = strip_accents(value).strip().lower()
    confidence_map = {
        "alta": "Alta",
        "media": "Média",
        "baixa": "Baixa",
    }
    return confidence_map.get(normalized, "Baixa")


def confidence_to_excel(value: str) -> str:
    normalized = strip_accents(value).strip().lower()
    if normalized == "alta":
        return "alta"
    if normalized == "media":
        return "media"
    return "baixa"


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = strip_accents(str(value)).strip().lower()
    return normalized in {"true", "1", "sim", "yes"}


def formato_json_for(competencia: str) -> str:
    return FORMATO_JSON_POR_COMPETENCIA.get(competencia, FORMATO_JSON).format(competencia=competencia)


def format_desvios_encontrados(raw: dict[str, Any]) -> str:
    desvios = raw.get("desvios_encontrados")
    if not isinstance(desvios, list):
        return ""

    evidencias: list[str] = []
    for desvio in desvios[:5]:
        if not isinstance(desvio, dict):
            continue
        tipo = str(desvio.get("tipo", "")).strip()
        trecho = str(desvio.get("trecho", "")).strip()
        gravidade = str(desvio.get("gravidade", "")).strip()
        partes = []
        if tipo:
            partes.append(tipo)
        if gravidade:
            partes.append(gravidade)
        label = "/".join(partes)
        if trecho and label:
            evidencias.append(f"{label}: {trecho}")
        elif trecho:
            evidencias.append(trecho)
        elif label:
            evidencias.append(label)

    return " | ".join(evidencias)


def format_repertorios_identificados(raw: dict[str, Any]) -> str:
    repertorios = raw.get("repertorios_identificados")
    if not isinstance(repertorios, list):
        return ""

    evidencias: list[str] = []
    for repertorio in repertorios[:5]:
        if not isinstance(repertorio, dict):
            continue
        referencia = str(repertorio.get("referencia", "")).strip()
        classificacao = str(repertorio.get("classificacao", "")).strip()
        justificativa = str(repertorio.get("justificativa", "")).strip()
        partes = []
        if referencia:
            partes.append(referencia)
        if classificacao:
            partes.append(f"classificação: {classificacao}")
        if justificativa:
            partes.append(justificativa)
        if partes:
            evidencias.append(" - ".join(partes))

    return " | ".join(evidencias)


def format_projeto_argumentacao(raw: dict[str, Any]) -> str:
    campos = {
        "projeto_de_texto": "Projeto",
        "tese_identificada": "Tese",
        "autoria": "Autoria",
    }
    evidencias = []
    for campo, label in campos.items():
        valor = raw.get(campo)
        if valor is None:
            continue
        texto = str(valor).strip()
        if texto:
            evidencias.append(f"{label}: {texto}")

    if raw.get("teto_por_tangenciamento_aplicado") is True:
        evidencias.append("Teto por tangenciamento aplicado")

    return " | ".join(evidencias)


def format_coesao_articulacao(raw: dict[str, Any]) -> str:
    evidencias = []
    articulacao = str(raw.get("articulacao_interparagrafo", "")).strip()
    diversidade = str(raw.get("diversidade_coesiva", "")).strip()
    if articulacao:
        evidencias.append(f"Articulação interparágrafo: {articulacao}")
    if diversidade:
        evidencias.append(f"Diversidade coesiva: {diversidade}")

    inadequacoes = raw.get("inadequacoes_identificadas")
    if isinstance(inadequacoes, list):
        for inadequacao in inadequacoes[:5]:
            if not isinstance(inadequacao, dict):
                continue
            tipo = str(inadequacao.get("tipo", "")).strip()
            trecho = str(inadequacao.get("trecho", "")).strip()
            if tipo and trecho:
                evidencias.append(f"{tipo}: {trecho}")
            elif tipo:
                evidencias.append(tipo)
            elif trecho:
                evidencias.append(trecho)

    return " | ".join(evidencias)


def format_proposta_intervencao(raw: dict[str, Any]) -> str:
    evidencias = []
    elementos = raw.get("elementos_identificados")
    if isinstance(elementos, dict):
        labels = {
            "agente": "Agente",
            "acao": "Ação",
            "meio": "Meio",
            "efeito": "Efeito",
            "detalhamento": "Detalhamento",
        }
        for campo, label in labels.items():
            valor = elementos.get(campo)
            if valor is None:
                continue
            texto = str(valor).strip()
            if texto and texto.lower() != "null":
                evidencias.append(f"{label}: {texto}")

    total = raw.get("total_elementos")
    if total is not None:
        evidencias.append(f"Total de elementos: {total}")

    if raw.get("articulacao_com_texto") is False:
        evidencias.append("Sem articulação suficiente com o texto")
    if raw.get("respeita_direitos_humanos") is False:
        evidencias.append("Desrespeito aos direitos humanos")
    if raw.get("teto_por_tangenciamento_aplicado") is True:
        evidencias.append("Teto por tangenciamento aplicado")

    return " | ".join(evidencias)


def build_gate_anulacao_prompt(
    tema: str,
    redacao: str,
    status_ocr: str,
    status_tema: str,
    num_linhas: int,
) -> str:
    return PROMPT_GATE_ANULACAO.format(
        tema=tema,
        status_tema=status_tema,
        status_ocr=status_ocr,
        num_linhas=num_linhas,
        redacao=redacao,
    ).strip()


def build_prompt(
    competencia: str,
    tema: str,
    redacao: str,
    status_ocr: str,
    rubrica: str,
    status_tema: str,
    status_anulacao: str,
    tangenciamento_c2: str,
) -> str:
    return PROMPT_BASE.format(
        competencia=competencia,
        tema=tema,
        status_tema=status_tema,
        status_ocr=status_ocr,
        status_anulacao=status_anulacao,
        tangenciamento_c2=tangenciamento_c2,
        rubrica=rubrica,
        redacao=redacao,
        formato_json=formato_json_for(competencia),
    ).strip()


def call_sabia(api_key: str, model: str, prompt: str, timeout: int, retries: int) -> SabiaResponse:
    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": prompt,
            }
        ],
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        request = urllib.request.Request(API_URL, data=data, headers=headers, method="POST")
        started_at = time.monotonic()
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                body = json.loads(response.read().decode("utf-8"))
            tokens_input, tokens_output, custo_estimado_usd = extract_sabia_usage(body)
            return SabiaResponse(
                content=body["choices"][0]["message"]["content"],
                response_body=body,
                tempo_resposta_s=round(time.monotonic() - started_at, 3),
                tokens_input=tokens_input,
                tokens_output=tokens_output,
                custo_estimado_usd=custo_estimado_usd,
            )
        except (urllib.error.URLError, urllib.error.HTTPError, KeyError, json.JSONDecodeError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise RuntimeError(f"Falha na chamada ao Sabiá: {exc}") from exc
    raise RuntimeError(f"Falha na chamada ao Sabiá: {last_error}")


def build_reprompt_with_validation_error(
    original_prompt: str,
    etapa: str,
    erro_validacao: str,
    json_resposta_bruto: str,
) -> str:
    return f"""{original_prompt}

CORREÇÃO OBRIGATÓRIA DO JSON
A resposta anterior da etapa {etapa} falhou na validação Pydantic do Corretor X.

Erro de validação:
{erro_validacao}

JSON anterior rejeitado:
{json_resposta_bruto}

Retorne novamente SOMENTE um JSON válido, sem markdown e sem texto fora do JSON.
Mantenha a avaliação pedagógica, mas corrija estritamente os campos, tipos e coerências apontados pelo erro.
""".strip()


def normalize_avaliacao(raw: dict[str, Any], competencia: str) -> AvaliacaoCompetencia:
    nota = int(raw.get("nota_corretor_x", raw.get("nota")))
    if nota not in VALID_SCORES:
        raise ValueError(f"Nota inválida para {competencia}: {nota}")
    nivel = normalize_confidence(str(raw.get("nivel_confianca", "")))
    evidencia = (
        str(raw.get("evidencia_no_texto", "")).strip()
        or format_desvios_encontrados(raw)
        or format_repertorios_identificados(raw)
        or format_projeto_argumentacao(raw)
        or format_coesao_articulacao(raw)
        or format_proposta_intervencao(raw)
    )
    return AvaliacaoCompetencia(
        competencia=competencia,
        nota_corretor_x=nota,
        comentario_do_erro=str(raw.get("comentario_do_erro", raw.get("comentario", ""))).strip(),
        evidencia_no_texto=evidencia,
        sugestao_de_melhoria=str(raw.get("sugestao_de_melhoria", raw.get("sugestao", ""))).strip(),
        nivel_confianca=nivel,
    )


def normalize_gate_anulacao(raw: dict[str, Any]) -> GateAnulacao:
    detalhes = raw.get("detalhes")
    motivos = raw.get("motivos")
    if not isinstance(motivos, list):
        motivos = []
    motivos = [str(motivo).strip() for motivo in motivos if str(motivo).strip()]

    if isinstance(detalhes, dict):
        motivos_por_detalhe = [
            str(motivo)
            for motivo, ativo in detalhes.items()
            if ativo is True and str(motivo) not in motivos
        ]
        motivos.extend(motivos_por_detalhe)

    anulado = bool(raw.get("anulado")) or bool(motivos)
    return GateAnulacao(
        anulado=anulado,
        motivos=motivos,
        evidencia=str(raw.get("evidencia", "")).strip(),
        nivel_confianca=normalize_confidence(str(raw.get("nivel_confianca", ""))),
    )


def validar_json_etapa(raw: dict[str, Any], etapa: str) -> tuple[Any | None, str]:
    try:
        schema = schema_for_etapa(etapa)
        objeto, erro = parse_resposta_sabia(raw, schema)
        if erro:
            return None, erro
        return objeto, ""
    except Exception as exc:
        return None, f"{type(exc).__name__}: {exc}"


def run_sabia_json_call_once(
    api_key: str,
    model: str,
    prompt: str,
    etapa: str,
    prompt_version: str,
    timeout: int,
    retries: int,
) -> tuple[dict[str, Any], ChamadaAuditoria, Any]:
    response = call_sabia(api_key, model, prompt, timeout, retries)
    validacao_ok = False
    erros_validacao = ""
    raw: dict[str, Any] = {}
    objeto_pydantic: Any | None = None
    json_resposta_bruto = response.content

    try:
        raw = extract_json(response.content)
        json_resposta_bruto = json.dumps(raw, ensure_ascii=False, sort_keys=True)
        objeto_pydantic, erros_validacao = validar_json_etapa(raw, etapa)
        validacao_ok = objeto_pydantic is not None
    except Exception as exc:
        erros_validacao = f"{type(exc).__name__}: {exc}"

    audit = ChamadaAuditoria(
        id_chamada=str(uuid.uuid4()),
        timestamp=datetime.now().replace(microsecond=0),
        etapa=etapa.lower(),
        modelo=model,
        prompt_versao=prompt_version,
        prompt_hash=prompt_hash(prompt),
        json_resposta_bruto=json_resposta_bruto,
        validacao_pydantic_ok=validacao_ok,
        erros_pydantic=erros_validacao,
        tempo_resposta_s=response.tempo_resposta_s,
        tokens_input=response.tokens_input,
        tokens_output=response.tokens_output,
        custo_estimado_usd=response.custo_estimado_usd,
    )

    if not validacao_ok:
        raise SabiaValidationError(etapa, erros_validacao, audit, raw)

    return raw, audit, objeto_pydantic


def run_sabia_json_call(
    api_key: str,
    model: str,
    fallback_model: str,
    prompt: str,
    etapa: str,
    prompt_version: str,
    timeout: int,
    retries: int,
) -> tuple[dict[str, Any], list[ChamadaAuditoria], Any]:
    chamadas: list[ChamadaAuditoria] = []
    last_error = ""
    last_json = ""

    recovery_strategies = [
        strategy
        for strategy in RETRY_POLICY["estrategias"]
        if strategy != "marcar_revisao_humana"
    ][: int(RETRY_POLICY["tentativas_max"])]
    attempts: list[tuple[str, str, str]] = [("chamada_inicial", model, prompt)]
    for strategy in recovery_strategies:
        attempt_model = (fallback_model or model) if strategy == "fallback_modelo_maior" else model
        attempts.append((strategy, attempt_model, ""))

    for index, (strategy, attempt_model, attempt_prompt) in enumerate(attempts):
        if index > 0:
            attempt_prompt = build_reprompt_with_validation_error(prompt, etapa, last_error, last_json)

        attempt_prompt_version = prompt_version if strategy == "chamada_inicial" else f"{prompt_version}:{strategy}"
        try:
            raw, audit, objeto_pydantic = run_sabia_json_call_once(
                api_key,
                attempt_model,
                attempt_prompt,
                etapa,
                attempt_prompt_version,
                timeout,
                retries,
            )
            chamadas.append(audit)
            return raw, chamadas, objeto_pydantic
        except SabiaValidationError as exc:
            last_error = exc.erro
            if exc.auditoria is not None:
                last_json = exc.auditoria.json_resposta_bruto
                chamadas.append(exc.auditoria)
            else:
                last_json = ""

    raise SabiaHumanReviewRequired(etapa, last_error, chamadas)


def total_tokens_auditoria(chamadas: list[ChamadaAuditoria]) -> int | None:
    total = 0
    found = False
    for chamada in chamadas:
        for value in (chamada.tokens_input, chamada.tokens_output):
            if value is not None:
                total += value
                found = True
    return total if found else None


def status_ocr_schema_value(status_ocr: str) -> str:
    normalized = strip_accents(status_ocr).strip().lower()
    if any(marker in normalized for marker in ("degrad", "baixa", "ilegivel", "incerta", "sem ocr")):
        return "degradado"
    if any(marker in normalized for marker in ("parcial", "manual", "revisao")):
        return "parcial"
    return "ok"


def build_metadados_redacao(
    case_id: str,
    aluno_nome: str,
    aluno_escola: str,
    tema: str,
    status_tema: str,
    status_ocr: str,
    num_linhas: int,
) -> MetadadosRedacao:
    return MetadadosRedacao(
        id_redacao=case_id,
        aluno_nome=aluno_nome,
        aluno_escola=aluno_escola,
        tema=tema,
        status_tema=status_tema,
        status_ocr=status_ocr_schema_value(status_ocr),
        num_linhas=num_linhas,
    )


def confianca_geral_pydantic(objetos: list[Any]) -> str:
    ranks = {"baixa": 0, "media": 1, "alta": 2}
    values: list[str] = []
    for objeto in objetos:
        value = getattr(objeto, "nivel_confianca", None)
        if value is None:
            continue
        values.append(str(getattr(value, "value", value)).lower())
    if not values:
        return "baixa"
    return min(values, key=lambda item: ranks.get(item, 0))


def respostas_anuladas_pydantic(gate: GateAnulacaoSchema) -> dict[str, Any]:
    nivel = str(getattr(gate.nivel_confianca, "value", gate.nivel_confianca))
    comentario = "Redação anulada antes da avaliação por competência."
    sugestao = "Resolver a condição de anulação antes de buscar pontuação por competência."
    abordagem_c2 = "fuga_total" if gate.detalhes.fuga_total_tema else "completa"
    tipo_textual_adequado = not gate.detalhes.nao_atendimento_tipo_textual
    return {
        "C1": RespostaC1(
            competencia="C1",
            nota=0,
            comentario=comentario,
            desvios_encontrados=[],
            total_desvios=0,
            reincidencia=False,
            tipo_reincidente=None,
            sugestao=sugestao,
            nivel_confianca=nivel,
        ),
        "C2": RespostaC2(
            competencia="C2",
            nota=0,
            comentario=comentario,
            abordagem_tema=abordagem_c2,
            tangenciamento_c2=False,
            alerta_tema=None,
            tipo_textual_adequado=tipo_textual_adequado,
            repertorios_identificados=[],
            sugestao=sugestao,
            nivel_confianca=nivel,
        ),
        "C3": RespostaC3(
            competencia="C3",
            nota=0,
            comentario=comentario,
            projeto_de_texto="ausente",
            tese_identificada=None,
            autoria="ausente",
            teto_por_tangenciamento_aplicado=False,
            sugestao=sugestao,
            nivel_confianca=nivel,
        ),
        "C4": RespostaC4(
            competencia="C4",
            nota=0,
            comentario=comentario,
            articulacao_interparagrafo="ausente",
            diversidade_coesiva="baixa",
            inadequacoes_identificadas=[],
            sugestao=sugestao,
            nivel_confianca=nivel,
        ),
        "C5": RespostaC5(
            competencia="C5",
            nota=0,
            comentario=comentario,
            elementos_identificados={},
            total_elementos=0,
            articulacao_com_texto=False,
            respeita_direitos_humanos=True,
            teto_por_tangenciamento_aplicado=False,
            sugestao=sugestao,
            nivel_confianca=nivel,
        ),
    }


def build_resultado_pydantic(
    metadados: MetadadosRedacao,
    resultado: ResultadoCorrecao,
    gate: GateAnulacaoSchema,
    respostas: dict[str, Any],
    alertas: list[str],
) -> ResultadoCorrecaoSchema:
    objetos = [respostas[competencia] for competencia in COMPETENCIAS] + [gate]
    return ResultadoCorrecaoSchema(
        metadados=metadados,
        anulada=resultado.anulado,
        gate=gate,
        tangenciamento=resultado.tangenciamento,
        c1=respostas["C1"],
        c2=respostas["C2"],
        c3=respostas["C3"],
        c4=respostas["C4"],
        c5=respostas["C5"],
        nota_final=resultado.nota_final,
        confianca_geral=confianca_geral_pydantic(objetos),
        alertas=alertas,
    )


def raw_indica_tangenciamento_c2(raw: dict[str, Any]) -> bool:
    abordagem = strip_accents(str(raw.get("abordagem_tema", ""))).strip().lower()
    return normalize_bool(raw.get("tangenciamento_c2")) or abordagem == "tangenciamento"


def append_sentence(text: str, sentence: str) -> str:
    text = text.strip()
    if sentence in text:
        return text
    if not text:
        return sentence
    return f"{text} {sentence}"


def append_evidence(text: str, evidence: str) -> str:
    text = text.strip()
    if evidence in text:
        return text
    if not text:
        return evidence
    return f"{text} | {evidence}"


def aplicar_teto_tangenciamento(avaliacao: AvaliacaoCompetencia) -> AvaliacaoCompetencia:
    if avaliacao.competencia not in COMPETENCIAS_COM_TETO_TANGENCIAMENTO:
        return avaliacao

    avaliacao.teto_tangenciamento_aplicado = True
    nota_original = avaliacao.nota_corretor_x
    if nota_original > 40:
        avaliacao.nota_corretor_x = 40
        avaliacao.comentario_do_erro = append_sentence(
            avaliacao.comentario_do_erro,
            f"Teto por tangenciamento aplicado: nota limitada de {nota_original} para 40.",
        )
    else:
        avaliacao.comentario_do_erro = append_sentence(
            avaliacao.comentario_do_erro,
            "Teto por tangenciamento aplicado; nota já estava dentro do limite.",
        )
    avaliacao.evidencia_no_texto = append_evidence(
        avaliacao.evidencia_no_texto,
        "Teto por tangenciamento aplicado",
    )
    return avaliacao


def avaliacoes_por_anulacao(gate: GateAnulacao) -> list[AvaliacaoCompetencia]:
    motivos = ", ".join(gate.motivos) if gate.motivos else "condição de anulação"
    comentario = f"Redação anulada antes da avaliação por competência. Motivo(s): {motivos}."
    evidencia = gate.evidencia or "Gate de anulação indicou anulação total."
    sugestao = "Resolver a condição de anulação antes de buscar pontuação por competência."
    return [
        AvaliacaoCompetencia(
            competencia=competencia,
            nota_corretor_x=0,
            comentario_do_erro=comentario,
            evidencia_no_texto=evidencia,
            sugestao_de_melhoria=sugestao,
            nivel_confianca=gate.nivel_confianca,
        )
        for competencia in COMPETENCIAS
    ]


def corrigir_competencia(
    competencia: str,
    api_key: str,
    model: str,
    fallback_model: str,
    prompt_version: str,
    tema: str,
    redacao: str,
    status_tema: str,
    status_ocr: str,
    status_anulacao: str,
    tangenciamento_c2: bool,
    prompts_dir: Path,
    timeout: int,
    retries: int,
) -> tuple[AvaliacaoCompetencia, dict[str, Any], list[ChamadaAuditoria], Any]:
    rubrica = load_competencia_prompt(competencia, prompts_dir)
    prompt = build_prompt(
        competencia,
        tema,
        redacao,
        status_ocr,
        rubrica,
        status_tema,
        status_anulacao,
        "true" if tangenciamento_c2 else "false",
    )
    raw, audits, resposta_pydantic = run_sabia_json_call(
        api_key,
        model,
        fallback_model,
        prompt,
        competencia.lower(),
        prompt_version,
        timeout,
        retries,
    )
    return normalize_avaliacao(raw, competencia), raw, audits, resposta_pydantic


def corrigir_redacao(
    api_key: str,
    model: str,
    fallback_model: str,
    prompt_version: str,
    tema: str,
    redacao: str,
    status_tema: str,
    status_ocr: str,
    status_anulacao: str,
    tangenciamento_c2_inicial: bool,
    num_linhas: int,
    prompts_dir: Path,
    timeout: int,
    retries: int,
    print_progress: bool = False,
    metadados: MetadadosRedacao | None = None,
) -> ResultadoCorrecao:
    started_at = time.monotonic()
    gate_prompt = build_gate_anulacao_prompt(tema, redacao, status_ocr, status_tema, num_linhas)
    gate_raw, gate_audits, gate_pydantic = run_sabia_json_call(
        api_key,
        model,
        fallback_model,
        gate_prompt,
        "gate",
        prompt_version,
        timeout,
        retries,
    )
    chamadas_auditoria = list(gate_audits)
    gate = normalize_gate_anulacao(gate_raw)
    if gate.anulado:
        avaliacoes = avaliacoes_por_anulacao(gate)
        nota_final = sum(avaliacao.nota_corretor_x for avaliacao in avaliacoes)
        if print_progress:
            print(f"GATE: redação anulada ({', '.join(gate.motivos)})")
        resultado = ResultadoCorrecao(
            anulado=True,
            motivos_anulacao=gate.motivos,
            tangenciamento=False,
            nota_final=nota_final,
            avaliacoes=avaliacoes,
            raw_respostas={"gate": gate_raw},
            chamadas_auditoria=chamadas_auditoria,
            custo_tokens_total=total_tokens_auditoria(chamadas_auditoria),
            tempo_processamento_s=round(time.monotonic() - started_at, 3),
        )
        if metadados is not None:
            respostas = respostas_anuladas_pydantic(gate_pydantic)
            alertas = build_alertas(status_tema, status_ocr, resultado).split("; ")
            alertas = [alerta for alerta in alertas if alerta]
            resultado.resultado_pydantic = build_resultado_pydantic(
                metadados,
                resultado,
                gate_pydantic,
                respostas,
                alertas,
            )
        return resultado

    if print_progress:
        print("GATE: nenhuma condição de anulação total detectada.")

    tangenciamento_c2_detectado = tangenciamento_c2_inicial
    avaliacoes_por_competencia: dict[str, AvaliacaoCompetencia] = {}
    raw_respostas: dict[str, dict[str, Any]] = {"gate": gate_raw}
    respostas_pydantic: dict[str, Any] = {}
    for competencia in ORDEM_CORRECAO:
        avaliacao, raw, audits, resposta_pydantic = corrigir_competencia(
            competencia,
            api_key,
            model,
            fallback_model,
            prompt_version,
            tema,
            redacao,
            status_tema,
            status_ocr,
            status_anulacao,
            tangenciamento_c2_detectado,
            prompts_dir,
            timeout,
            retries,
        )
        chamadas_auditoria.extend(audits)
        raw_respostas[competencia] = raw
        if competencia == "C2" and raw_indica_tangenciamento_c2(raw):
            tangenciamento_c2_detectado = True
        if tangenciamento_c2_detectado and competencia in COMPETENCIAS_COM_TETO_TANGENCIAMENTO:
            avaliacao = aplicar_teto_tangenciamento(avaliacao)
            resposta_pydantic = resposta_pydantic.model_copy(
                update={
                    "nota": min(resposta_pydantic.nota, 40),
                    "teto_por_tangenciamento_aplicado": True,
                }
            )
        respostas_pydantic[competencia] = resposta_pydantic
        avaliacoes_por_competencia[competencia] = avaliacao
        if print_progress:
            print(f"{competencia}: {avaliacao.nota_corretor_x}/200 ({avaliacao.nivel_confianca})")

    avaliacoes = [avaliacoes_por_competencia[competencia] for competencia in COMPETENCIAS]
    nota_final = sum(avaliacao.nota_corretor_x for avaliacao in avaliacoes)
    resultado = ResultadoCorrecao(
        anulado=False,
        motivos_anulacao=[],
        tangenciamento=tangenciamento_c2_detectado,
        nota_final=nota_final,
        avaliacoes=avaliacoes,
        raw_respostas=raw_respostas,
        chamadas_auditoria=chamadas_auditoria,
        custo_tokens_total=total_tokens_auditoria(chamadas_auditoria),
        tempo_processamento_s=round(time.monotonic() - started_at, 3),
    )
    if metadados is not None:
        alertas = build_alertas(status_tema, status_ocr, resultado).split("; ")
        alertas = [alerta for alerta in alertas if alerta]
        resultado.resultado_pydantic = build_resultado_pydantic(
            metadados,
            resultado,
            gate_pydantic,
            respostas_pydantic,
            alertas,
        )
    return resultado


def avaliacoes_por_competencia(avaliacoes: list[AvaliacaoCompetencia]) -> dict[str, AvaliacaoCompetencia]:
    return {avaliacao.competencia.lower(): avaliacao for avaliacao in avaliacoes}


def confianca_geral(avaliacoes: list[AvaliacaoCompetencia]) -> str:
    ranks = {"baixa": 0, "media": 1, "alta": 2}
    confidences = [confidence_to_excel(avaliacao.nivel_confianca) for avaliacao in avaliacoes]
    if not confidences:
        return "baixa"
    return min(confidences, key=lambda item: ranks.get(item, 0))


def nivel_proficiencia(nota_final: int) -> str:
    if nota_final < 500:
        return "Insuficiente"
    if nota_final < 700:
        return "Regular"
    if nota_final < 850:
        return "Bom"
    return "Excelente"


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def raw_competencia(resultado: ResultadoCorrecao, competencia: str) -> dict[str, Any]:
    raw = resultado.raw_respostas.get(competencia.upper(), {})
    return raw if isinstance(raw, dict) else {}


def count_repertorios(raw_c2: dict[str, Any]) -> int:
    repertorios = raw_c2.get("repertorios_identificados")
    return len(repertorios) if isinstance(repertorios, list) else 0


def count_repertorios_bolso(raw_c2: dict[str, Any]) -> int:
    repertorios = raw_c2.get("repertorios_identificados")
    if not isinstance(repertorios, list):
        return 0
    return sum(
        1
        for repertorio in repertorios
        if isinstance(repertorio, dict)
        and strip_accents(str(repertorio.get("classificacao", ""))).strip().lower() == "bolso"
    )


def text_or_blank(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if strip_accents(text).lower() in {"", "null", "none", "n/a"}:
        return ""
    return text


def slugify_calibracao(value: Any) -> str:
    text = strip_accents(text_or_blank(value)).lower()
    tokens = re.findall(r"[a-z0-9]+", text)
    return "_".join(tokens)


def repertorio_subtipo(referencia: Any) -> str:
    text = text_or_blank(referencia)
    if not text:
        return "repertorio"

    normalized_text = strip_accents(text).lower()
    for alias, slug in REPERTORIO_SUBTIPO_ALIASES.items():
        if alias in normalized_text:
            return slug

    if " - " in text:
        first_segment = text.split(" - ", 1)[0]
        slug = slugify_calibracao(first_segment)
        if slug:
            return slug

    match = re.match(r"(.+?),\s+d[eoas]*\s+(.+)$", text, flags=re.IGNORECASE)
    if match:
        title = slugify_calibracao(match.group(1))
        author_tokens = re.findall(r"[a-z0-9]+", strip_accents(match.group(2)).lower())
        if title and author_tokens:
            return f"{title}_{author_tokens[-1]}"

    return slugify_calibracao(text) or "repertorio"


def build_alertas(status_tema: str, status_ocr: str, resultado: ResultadoCorrecao) -> str:
    alertas: list[str] = []
    status_tema_normalizado = strip_accents(status_tema).strip().lower()
    if status_tema_normalizado == "inferido":
        alertas.append("tema_inferido")
    elif status_tema_normalizado == "ausente":
        alertas.append("tema_ausente")

    status_ocr_normalizado = strip_accents(status_ocr).strip().lower()
    if any(
        marker in status_ocr_normalizado
        for marker in (
            "baixa",
            "ilegivel",
            "incerta",
            "manual",
            "sem ocr",
            "revisao humana",
            "parcial",
            "ocr seguro",
            "nao valid",
            "não valid",
            "trecho critico",
            "trecho crítico",
        )
    ):
        alertas.append("ocr_revisao_humana_recomendada")

    if resultado.anulado:
        alertas.append("redacao_anulada")
    if resultado.tangenciamento:
        alertas.append("tangenciamento")

    por_competencia = avaliacoes_por_competencia(resultado.avaliacoes)
    if por_competencia.get("c3") and por_competencia["c3"].teto_tangenciamento_aplicado:
        alertas.append("teto_aplicado_c3")
    if por_competencia.get("c5") and por_competencia["c5"].teto_tangenciamento_aplicado:
        alertas.append("teto_aplicado_c5")

    return "; ".join(dict.fromkeys(alertas))


def is_unsafe_ocr_status(status_ocr: str) -> bool:
    """Return True when the transcript cannot be trusted for ENEM scoring."""
    normalized = strip_accents(status_ocr).strip().lower()
    if not normalized:
        return False

    safe_prefixes = (
        "ok:",
        "validado:",
        "literal_validada:",
        "literal validada:",
        "nao necessario",
        "não necessario",
        "nao necessário",
        "não necessário",
    )
    if normalized.startswith(safe_prefixes):
        return False

    unsafe_markers = (
        "aguardando_ocr",
        "ocr_degradado",
        "degradado",
        "parcial",
        "ocr seguro",
        "baixa",
        "ilegivel",
        "incerta",
        "manual assistida",
        "sem ocr",
        "revisao",
        "revisar",
        "pendente",
        "nao valid",
        "não valid",
        "trecho critico",
        "trecho crítico",
        "safe_for_correction=false",
    )
    return any(marker in normalized for marker in unsafe_markers)


def build_ocr_unsafe_alertas(status_tema: str) -> str:
    alertas: list[str] = []
    status_tema_normalizado = strip_accents(status_tema).strip().lower()
    if status_tema_normalizado == "inferido":
        alertas.append("tema_inferido")
    elif status_tema_normalizado == "ausente":
        alertas.append("tema_ausente")
    alertas.extend(
        [
            "ocr_nao_validado",
            "correcao_bloqueada_por_ocr",
            "c1_nao_auditavel_sem_transcricao_literal",
        ]
    )
    return "; ".join(dict.fromkeys(alertas))


def reset_worksheet(ws: Any) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def style_devolutiva(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1D1D1F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 18,
        "B": 20,
        "C": 24,
        "D": 24,
        "E": 42,
        "F": 14,
        "G": 10,
        "H": 28,
        "I": 16,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 12,
        "P": 16,
        "Q": 16,
        "R": 42,
        "S": 42,
        "T": 42,
        "U": 42,
        "V": 42,
        "W": 34,
        "X": 34,
        "Y": 34,
        "Z": 34,
        "AA": 34,
        "AB": 16,
        "AC": 36,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 36
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 90

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DEVOLUTIVA_HEADERS))}{max(ws.max_row, 1)}"


def style_resumo(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="007AFF")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 18,
        "B": 20,
        "C": 24,
        "D": 42,
        "E": 12,
        "F": 18,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 16,
        "M": 16,
        "N": 20,
        "O": 16,
        "P": 20,
        "Q": 20,
        "R": 16,
        "S": 20,
        "T": 18,
        "U": 22,
        "V": 16,
        "W": 16,
        "X": 18,
        "Y": 22,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 36
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RESUMO_HEADERS))}{max(ws.max_row, 1)}"


def style_calibracao(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="34C759")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 18,
        "B": 14,
        "C": 24,
        "D": 24,
        "E": 54,
        "F": 14,
        "G": 18,
        "H": 54,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 36
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 54

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CALIBRACAO_HEADERS))}{max(ws.max_row, 1)}"


def style_auditoria(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="FF9500")
    header_font = Font(color="1D1D1F", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 38,
        "B": 18,
        "C": 20,
        "D": 10,
        "E": 14,
        "F": 14,
        "G": 68,
        "H": 80,
        "I": 20,
        "J": 36,
        "K": 18,
        "L": 14,
        "M": 14,
        "N": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 36
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 96

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(AUDITORIA_HEADERS))}{max(ws.max_row, 1)}"


def build_resumo_row(
    case_id: str,
    data_correcao: datetime,
    aluno_escola: str,
    tema: str,
    resultado: ResultadoCorrecao,
) -> list[Any]:
    por_competencia = avaliacoes_por_competencia(resultado.avaliacoes)
    raw_c1 = raw_competencia(resultado, "C1")
    raw_c2 = raw_competencia(resultado, "C2")
    raw_c3 = raw_competencia(resultado, "C3")
    raw_c4 = raw_competencia(resultado, "C4")
    raw_c5 = raw_competencia(resultado, "C5")

    total_desvios_c1 = safe_int(raw_c1.get("total_desvios"))
    if total_desvios_c1 == 0 and isinstance(raw_c1.get("desvios_encontrados"), list):
        total_desvios_c1 = len(raw_c1["desvios_encontrados"])

    return [
        case_id,
        data_correcao,
        aluno_escola,
        tema,
        resultado.nota_final,
        nivel_proficiencia(resultado.nota_final),
        por_competencia["c1"].nota_corretor_x,
        por_competencia["c2"].nota_corretor_x,
        por_competencia["c3"].nota_corretor_x,
        por_competencia["c4"].nota_corretor_x,
        por_competencia["c5"].nota_corretor_x,
        total_desvios_c1,
        normalize_bool(raw_c1.get("reincidencia")),
        str(raw_c2.get("abordagem_tema", "")).strip(),
        count_repertorios(raw_c2),
        count_repertorios_bolso(raw_c2),
        str(raw_c3.get("projeto_de_texto", "")).strip(),
        str(raw_c3.get("autoria", "")).strip(),
        str(raw_c4.get("diversidade_coesiva", "")).strip(),
        safe_int(raw_c5.get("total_elementos")),
        normalize_bool(raw_c5.get("articulacao_com_texto")),
        normalize_bool(raw_c5.get("respeita_direitos_humanos")),
        confianca_geral(resultado.avaliacoes),
        resultado.custo_tokens_total,
        resultado.tempo_processamento_s,
    ]


def append_calibracao_c1(rows: list[list[str]], case_id: str, raw_c1: dict[str, Any]) -> None:
    desvios = raw_c1.get("desvios_encontrados")
    if not isinstance(desvios, list):
        return

    for desvio in desvios:
        if not isinstance(desvio, dict):
            continue
        rows.append(
            [
                case_id,
                "C1",
                "desvio",
                text_or_blank(desvio.get("tipo")),
                text_or_blank(desvio.get("trecho")),
                text_or_blank(desvio.get("gravidade")),
                "",
                "",
            ]
        )


def append_calibracao_c2(rows: list[list[str]], case_id: str, raw_c2: dict[str, Any]) -> None:
    repertorios = raw_c2.get("repertorios_identificados")
    if not isinstance(repertorios, list):
        return

    for repertorio in repertorios:
        if not isinstance(repertorio, dict):
            continue
        referencia = text_or_blank(repertorio.get("referencia"))
        classificacao = text_or_blank(repertorio.get("classificacao"))
        rows.append(
            [
                case_id,
                "C2",
                "repertorio",
                repertorio_subtipo(referencia),
                referencia,
                "",
                classificacao,
                text_or_blank(repertorio.get("justificativa")),
            ]
        )


def append_calibracao_c4(rows: list[list[str]], case_id: str, raw_c4: dict[str, Any]) -> None:
    inadequacoes = raw_c4.get("inadequacoes_identificadas")
    if not isinstance(inadequacoes, list):
        return

    for inadequacao in inadequacoes:
        if not isinstance(inadequacao, dict):
            continue
        rows.append(
            [
                case_id,
                "C4",
                "inadequacao_coesiva",
                text_or_blank(inadequacao.get("tipo")),
                text_or_blank(inadequacao.get("trecho")),
                "",
                "",
                text_or_blank(inadequacao.get("justificativa")),
            ]
        )


def append_calibracao_c5(rows: list[list[str]], case_id: str, raw_c5: dict[str, Any]) -> None:
    elementos = raw_c5.get("elementos_identificados")
    if not isinstance(elementos, dict):
        return

    for subtipo in ("agente", "acao", "meio", "efeito", "detalhamento"):
        trecho = text_or_blank(elementos.get(subtipo))
        rows.append(
            [
                case_id,
                "C5",
                "elemento_proposta",
                subtipo,
                trecho,
                "",
                "",
                f"{subtipo} identificado" if trecho else f"{subtipo} ausente",
            ]
        )


def build_calibracao_rows(case_id: str, resultado: ResultadoCorrecao) -> list[list[str]]:
    rows: list[list[str]] = []
    append_calibracao_c1(rows, case_id, raw_competencia(resultado, "C1"))
    append_calibracao_c2(rows, case_id, raw_competencia(resultado, "C2"))
    append_calibracao_c4(rows, case_id, raw_competencia(resultado, "C4"))
    append_calibracao_c5(rows, case_id, raw_competencia(resultado, "C5"))
    return rows


def build_auditoria_rows(case_id: str, resultado: ResultadoCorrecao) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for chamada in resultado.chamadas_auditoria:
        rows.append(
            [
                chamada.id_chamada,
                case_id,
                chamada.timestamp,
                chamada.etapa,
                chamada.modelo,
                chamada.prompt_versao,
                chamada.prompt_hash,
                chamada.json_resposta_bruto,
                chamada.validacao_pydantic_ok,
                chamada.erros_pydantic,
                chamada.tempo_resposta_s,
                chamada.tokens_input,
                chamada.tokens_output,
                chamada.custo_estimado_usd,
            ]
        )
    return rows


def fill_workbook(
    template_path: Path,
    output_path: Path,
    case_id: str,
    aluno_nome: str,
    aluno_escola: str,
    tema: str,
    status_tema: str,
    status_ocr: str,
    resultado: ResultadoCorrecao,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    data_correcao = datetime.now().replace(microsecond=0)

    ws = wb["Devolutiva"] if "Devolutiva" in wb.sheetnames else wb.create_sheet("Devolutiva")
    reset_worksheet(ws)

    por_competencia = avaliacoes_por_competencia(resultado.avaliacoes)
    def avaliacao(competencia: str) -> AvaliacaoCompetencia:
        return por_competencia[competencia.lower()]

    row = [
        case_id,
        data_correcao,
        aluno_nome,
        aluno_escola,
        tema,
        status_tema,
        resultado.anulado,
        "; ".join(resultado.motivos_anulacao),
        resultado.tangenciamento,
        avaliacao("C1").nota_corretor_x,
        avaliacao("C2").nota_corretor_x,
        avaliacao("C3").nota_corretor_x,
        avaliacao("C4").nota_corretor_x,
        avaliacao("C5").nota_corretor_x,
        resultado.nota_final,
        avaliacao("C3").teto_tangenciamento_aplicado,
        avaliacao("C5").teto_tangenciamento_aplicado,
        avaliacao("C1").comentario_do_erro,
        avaliacao("C2").comentario_do_erro,
        avaliacao("C3").comentario_do_erro,
        avaliacao("C4").comentario_do_erro,
        avaliacao("C5").comentario_do_erro,
        avaliacao("C1").sugestao_de_melhoria,
        avaliacao("C2").sugestao_de_melhoria,
        avaliacao("C3").sugestao_de_melhoria,
        avaliacao("C4").sugestao_de_melhoria,
        avaliacao("C5").sugestao_de_melhoria,
        confianca_geral(resultado.avaliacoes),
        build_alertas(status_tema, status_ocr, resultado),
    ]
    ws.append(DEVOLUTIVA_HEADERS)
    ws.append(row)
    ws["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    style_devolutiva(ws)

    ws_resumo = wb["Resumo"] if "Resumo" in wb.sheetnames else wb.create_sheet("Resumo")
    reset_worksheet(ws_resumo)
    ws_resumo.append(RESUMO_HEADERS)
    ws_resumo.append(build_resumo_row(case_id, data_correcao, aluno_escola, tema, resultado))
    ws_resumo["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws_resumo["Y2"].number_format = "0.000"
    style_resumo(ws_resumo)

    ws_calibracao = wb["Calibração"] if "Calibração" in wb.sheetnames else wb.create_sheet("Calibração")
    reset_worksheet(ws_calibracao)
    ws_calibracao.append(CALIBRACAO_HEADERS)
    for calibracao_row in build_calibracao_rows(case_id, resultado):
        ws_calibracao.append(calibracao_row)
    style_calibracao(ws_calibracao)

    ws_auditoria = wb["Auditoria"] if "Auditoria" in wb.sheetnames else wb.create_sheet("Auditoria")
    reset_worksheet(ws_auditoria)
    ws_auditoria.append(AUDITORIA_HEADERS)
    for auditoria_row in build_auditoria_rows(case_id, resultado):
        ws_auditoria.append(auditoria_row)
    for row_idx in range(2, ws_auditoria.max_row + 1):
        ws_auditoria[f"C{row_idx}"].number_format = "yyyy-mm-dd hh:mm:ss"
        ws_auditoria[f"K{row_idx}"].number_format = "0.000"
        ws_auditoria[f"N{row_idx}"].number_format = "0.000000"
    style_auditoria(ws_auditoria)

    wb.save(output_path)


def fill_ocr_unsafe_workbook(
    template_path: Path,
    output_path: Path,
    case_id: str,
    aluno_nome: str,
    aluno_escola: str,
    tema: str,
    status_tema: str,
    status_ocr: str,
    redacao: str,
    num_linhas: int,
) -> None:
    """Generate a non-scored workbook when OCR is not safe enough for grading."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    data_correcao = datetime.now().replace(microsecond=0)

    comentario_bloqueio = (
        "Correção por competência não executada: a transcrição OCR não foi validada "
        "como literal. Uma nota de C1 baseada nesse texto poderia confundir erros reais "
        "do aluno com erros de leitura da imagem."
    )
    comentario_demais = (
        "Correção não executada porque a base textual está insegura. Repertório, "
        "argumentação, coesão e proposta também podem ser distorcidos por OCR parcial."
    )
    sugestao = (
        "Reprocessar a imagem com maior resolução, melhor enquadramento e boa iluminação; "
        "depois rodar novamente para gerar notas."
    )
    alertas = build_ocr_unsafe_alertas(status_tema)

    ws = wb["Devolutiva"] if "Devolutiva" in wb.sheetnames else wb.create_sheet("Devolutiva")
    reset_worksheet(ws)
    ws.append(DEVOLUTIVA_HEADERS)
    ws.append(
        [
            case_id,
            data_correcao,
            aluno_nome,
            aluno_escola,
            tema,
            status_tema,
            False,
            "",
            False,
            None,
            None,
            None,
            None,
            None,
            None,
            False,
            False,
            comentario_bloqueio,
            comentario_demais,
            comentario_demais,
            comentario_demais,
            comentario_demais,
            sugestao,
            sugestao,
            sugestao,
            sugestao,
            sugestao,
            "baixa",
            alertas,
        ]
    )
    ws["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    style_devolutiva(ws)

    ws_resumo = wb["Resumo"] if "Resumo" in wb.sheetnames else wb.create_sheet("Resumo")
    reset_worksheet(ws_resumo)
    ws_resumo.append(RESUMO_HEADERS)
    ws_resumo.append(
        [
            case_id,
            data_correcao,
            aluno_escola,
            tema,
            None,
            "Não calculado",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "",
            None,
            None,
            "",
            "",
            "",
            None,
            None,
            None,
            "baixa",
            None,
            None,
        ]
    )
    ws_resumo["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    style_resumo(ws_resumo)

    ws_calibracao = wb["Calibração"] if "Calibração" in wb.sheetnames else wb.create_sheet("Calibração")
    reset_worksheet(ws_calibracao)
    ws_calibracao.append(CALIBRACAO_HEADERS)
    ws_calibracao.append(
        [
            case_id,
            "OCR",
            "bloqueio_correcao",
            "transcricao_nao_validada",
            status_ocr,
            "grave",
            "",
            "A transcrição não é segura para distinguir erro do aluno de erro de OCR.",
        ]
    )
    style_calibracao(ws_calibracao)

    ws_auditoria = wb["Auditoria"] if "Auditoria" in wb.sheetnames else wb.create_sheet("Auditoria")
    reset_worksheet(ws_auditoria)
    ws_auditoria.append(AUDITORIA_HEADERS)
    payload = {
        "acao": "correcao_bloqueada_por_ocr",
        "status_ocr": status_ocr,
        "status_tema": status_tema,
        "num_linhas_estimadas": num_linhas,
        "caracteres_transcricao": len(redacao),
        "motivo": "transcricao_nao_validada_como_literal",
    }
    ws_auditoria.append(
        [
            str(uuid.uuid4()),
            case_id,
            data_correcao,
            "ocr_gate",
            "local",
            DEFAULT_PROMPT_VERSION,
            prompt_hash(f"ocr_gate:{status_ocr}"),
            json.dumps(payload, ensure_ascii=False),
            True,
            "",
            0.0,
            0,
            0,
            0.0,
        ]
    )
    ws_auditoria["C2"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws_auditoria["K2"].number_format = "0.000"
    ws_auditoria["N2"].number_format = "0.000000"
    style_auditoria(ws_auditoria)

    wb.save(output_path)


def write_revisao_humana_marker(
    output_path: Path,
    case_id: str,
    exc: SabiaHumanReviewRequired,
) -> Path:
    marker_path = output_path.with_suffix(".revisao-humana.txt")
    marker_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "revisao_humana_recomendada=true",
        f"id_redacao={case_id}",
        f"etapa={exc.etapa}",
        f"erro={exc.erro}",
        f"tentativas={len(exc.chamadas_auditoria)}",
        "",
        "auditoria_tentativas:",
    ]
    for chamada in exc.chamadas_auditoria:
        lines.append(
            " | ".join(
                [
                    f"id_chamada={chamada.id_chamada}",
                    f"etapa={chamada.etapa}",
                    f"modelo={chamada.modelo}",
                    f"prompt_versao={chamada.prompt_versao}",
                    f"validacao_pydantic_ok={chamada.validacao_pydantic_ok}",
                    f"erro={chamada.erros_pydantic}",
                ]
            )
        )
    marker_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return marker_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Corrige redação ENEM com Sabiá e preenche Excel.")
    parser.add_argument("--tema-file", required=True, type=Path)
    parser.add_argument("--redacao-file", required=True, type=Path)
    parser.add_argument("--case-id", default="CASO-001")
    parser.add_argument("--aluno-id", default="Aluno 001")
    parser.add_argument("--aluno-nome")
    parser.add_argument("--aluno-escola", default="")
    parser.add_argument("--status-ocr", default="Não necessário")
    parser.add_argument("--status-tema", choices=["verificado", "inferido", "ausente"])
    parser.add_argument("--status-anulacao", default="nenhuma")
    parser.add_argument("--tangenciamento-c2", action="store_true")
    parser.add_argument("--num-linhas", type=int)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--fallback-model", default=DEFAULT_FALLBACK_MODEL)
    parser.add_argument("--prompt-version", default=DEFAULT_PROMPT_VERSION)
    parser.add_argument("--timeout", default=90, type=int)
    parser.add_argument("--retries", default=1, type=int)
    parser.add_argument("--dry-run", action="store_true", help="Não chama API; só valida entradas.")
    parser.add_argument(
        "--allow-unsafe-ocr-scoring",
        action="store_true",
        help=(
            "Força correção mesmo com OCR parcial/degradado. Use apenas para auditoria, "
            "pois C1 pode confundir erro do aluno com erro de leitura."
        ),
    )
    parser.add_argument(
        "--brain-prompts-dir",
        type=Path,
        default=Path("app-config/prompts"),
        help="Diretório do vault com prompts/rubricas por competência.",
    )
    parser.add_argument(
        "--template-xlsx",
        type=Path,
        default=Path("Cérebro do 1000/templates/Template - Entrega Excel Corretor X.xlsx"),
    )
    parser.add_argument(
        "--out-xlsx",
        type=Path,
        default=Path("Cérebro do 1000/casos/exports/CASO-001.xlsx"),
    )
    args = parser.parse_args()

    tema = read_text_or_default(args.tema_file, "Tema oficial não informado.")
    redacao = read_text(args.redacao_file)
    status_tema = args.status_tema or ("ausente" if tema == "Tema oficial não informado." else infer_status_tema(tema))
    num_linhas = args.num_linhas if args.num_linhas is not None else estimate_num_linhas(redacao)
    tangenciamento_c2_detectado = args.tangenciamento_c2
    ocr_inseguro = is_unsafe_ocr_status(args.status_ocr)
    aluno_nome = args.aluno_nome or args.aluno_id

    if args.dry_run:
        print("Entradas válidas.")
        print(f"Tema: {len(tema)} caracteres")
        print(f"Status do tema: {status_tema}")
        print(f"Status de anulação: {args.status_anulacao}")
        print(f"OCR seguro para nota: {'false' if ocr_inseguro else 'true'}")
        print(f"Número de linhas estimadas: {num_linhas}")
        print(f"Tangenciamento C2: {'true' if tangenciamento_c2_detectado else 'false'}")
        print(f"Redação: {len(redacao)} caracteres")
        print("Dry-run: nenhuma chamada ao Sabiá foi feita.")
        return 0

    if ocr_inseguro and not args.allow_unsafe_ocr_scoring:
        fill_ocr_unsafe_workbook(
            args.template_xlsx,
            args.out_xlsx,
            args.case_id,
            aluno_nome,
            args.aluno_escola,
            tema,
            status_tema,
            args.status_ocr,
            redacao,
            num_linhas,
        )
        print("Correção ENEM bloqueada: OCR/transcrição não validada como literal.")
        print("Notas não calculadas para evitar C1 e nota final contaminadas por erro de leitura.")
        print(f"Status OCR: {args.status_ocr}")
        print(f"Excel salvo em: {args.out_xlsx}")
        return 0

    api_key = os.environ.get("SABIA_API_KEY")
    if not api_key:
        print("Erro: defina SABIA_API_KEY no ambiente. Nunca coloque a chave no repositório.", file=sys.stderr)
        return 2

    metadados = build_metadados_redacao(
        args.case_id,
        aluno_nome,
        args.aluno_escola,
        tema,
        status_tema,
        args.status_ocr,
        num_linhas,
    )

    try:
        resultado = corrigir_redacao(
            api_key=api_key,
            model=args.model,
            fallback_model=args.fallback_model,
            prompt_version=args.prompt_version,
            tema=tema,
            redacao=redacao,
            status_tema=status_tema,
            status_ocr=args.status_ocr,
            status_anulacao=args.status_anulacao,
            tangenciamento_c2_inicial=tangenciamento_c2_detectado,
            num_linhas=num_linhas,
            prompts_dir=args.brain_prompts_dir,
            timeout=args.timeout,
            retries=args.retries,
            print_progress=True,
            metadados=metadados,
        )
    except SabiaHumanReviewRequired as exc:
        marker_path = write_revisao_humana_marker(args.out_xlsx, args.case_id, exc)
        print(
            f"Revisão humana recomendada: etapa {exc.etapa} falhou após "
            f"{len(exc.chamadas_auditoria)} tentativa(s).",
            file=sys.stderr,
        )
        print(f"Marcador salvo em: {marker_path}", file=sys.stderr)
        return REVISAO_HUMANA_EXIT_CODE

    fill_workbook(
        args.template_xlsx,
        args.out_xlsx,
        args.case_id,
        aluno_nome,
        args.aluno_escola,
        tema,
        status_tema,
        args.status_ocr,
        resultado,
    )
    if not resultado.anulado:
        print(f"Tangenciamento C2: {'true' if resultado.tangenciamento else 'false'}")
    print(f"Nota final estimada: {resultado.nota_final}/1000")
    print(f"Excel salvo em: {args.out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
